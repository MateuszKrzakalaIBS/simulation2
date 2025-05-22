"""
External Data Integration module for demographic simulation.
This module provides functions to download and integrate data from external sources
like Eurostat, World Bank, or local statistical offices.
"""

import pandas as pd
import numpy as np
import requests
import os
import json
import time
import re
from openpyxl import load_workbook
from io import BytesIO


class DataIntegrator:
    """Class for integrating external data with simulation inputs."""
    
    def __init__(self, input_file="Input.xlsx", cache_folder="data_cache"):
        """
        Initialize the DataIntegrator.
        
        Parameters:
        -----------
        input_file : str
            Path to the input Excel file
        cache_folder : str
            Folder to cache downloaded data
        """
        self.input_file = input_file
        self.cache_folder = cache_folder
        self.cache_index_file = os.path.join(cache_folder, "cache_index.json")
        self.eurostat_base_url = "https://ec.europa.eu/eurostat/api/dissemination/sdmx/2.1/data/"
        self.worldbank_base_url = "https://api.worldbank.org/v2/"
        
        # Ensure cache folder exists
        os.makedirs(cache_folder, exist_ok=True)
        
        # Initialize cache index if it doesn't exist
        if not os.path.exists(self.cache_index_file):
            with open(self.cache_index_file, 'w') as f:
                json.dump({}, f)
    
    def _get_cached_data(self, source, dataset_id, params=None):
        """
        Get data from cache if available and not expired.
        
        Parameters:
        -----------
        source : str
            Data source identifier (e.g., 'eurostat', 'worldbank')
        dataset_id : str
            Dataset identifier
        params : dict
            Parameters used for the request
            
        Returns:
        --------
        pandas.DataFrame or None
            Cached data if available, None otherwise
        """
        try:
            # Generate a cache key based on source, dataset, and params
            cache_key = f"{source}_{dataset_id}"
            if params:
                cache_key += "_" + "_".join(f"{k}_{v}" for k, v in sorted(params.items()))
            
            # Load cache index
            with open(self.cache_index_file, 'r') as f:
                cache_index = json.load(f)
            
            # Check if data is in cache and not expired (older than 30 days)
            if cache_key in cache_index:
                cache_entry = cache_index[cache_key]
                cache_file = cache_entry.get('file')
                
                if os.path.exists(os.path.join(self.cache_folder, cache_file)):
                    # Calculate age of cache in days
                    cache_age_days = (time.time() - cache_entry.get('timestamp', 0)) / (60 * 60 * 24)
                    
                    # If cache is less than 30 days old, use it
                    if cache_age_days < 30:
                        print(f"Using cached data for {dataset_id} (age: {cache_age_days:.1f} days)")
                        return pd.read_pickle(os.path.join(self.cache_folder, cache_file))
        except Exception as e:
            print(f"Error reading cache: {str(e)}")
        
        return None
    
    def _save_to_cache(self, df, source, dataset_id, params=None):
        """
        Save data to cache.
        
        Parameters:
        -----------
        df : pandas.DataFrame
            Data to cache
        source : str
            Data source identifier
        dataset_id : str
            Dataset identifier
        params : dict
            Parameters used for the request
            
        Returns:
        --------
        str
            Cache key
        """
        try:
            # Generate a cache key
            cache_key = f"{source}_{dataset_id}"
            if params:
                cache_key += "_" + "_".join(f"{k}_{v}" for k, v in sorted(params.items()))
            
            # Generate a filename
            filename = re.sub(r'[^a-zA-Z0-9_]', '_', cache_key) + '.pkl'
            
            # Save the DataFrame
            df.to_pickle(os.path.join(self.cache_folder, filename))
            
            # Update cache index
            with open(self.cache_index_file, 'r') as f:
                cache_index = json.load(f)
            
            cache_index[cache_key] = {
                'file': filename,
                'timestamp': time.time(),
                'source': source,
                'dataset_id': dataset_id,
                'params': params
            }
            
            with open(self.cache_index_file, 'w') as f:
                json.dump(cache_index, f, indent=2)
            
            print(f"Cached data for {dataset_id}")
            return cache_key
            
        except Exception as e:
            print(f"Error saving to cache: {str(e)}")
            return None
    
    def get_eurostat_data(self, dataset_id, dimension_filters=None, use_cache=True):
        """
        Get data from Eurostat.
        
        Parameters:
        -----------
        dataset_id : str
            Eurostat dataset identifier (e.g., 'demo_pjan')
        dimension_filters : dict
            Dictionary of dimension filters (e.g., {'geo': 'PL', 'sex': 'T'})
        use_cache : bool
            Whether to use cached data if available
            
        Returns:
        --------
        pandas.DataFrame
            DataFrame with Eurostat data
        """
        # Try to get from cache if enabled
        if use_cache:
            cached_data = self._get_cached_data('eurostat', dataset_id, dimension_filters)
            if cached_data is not None:
                return cached_data
        
        # Construct URL and parameters
        url = f"{self.eurostat_base_url}{dataset_id}/all/all"
        params = {
            'format': 'JSON',
            'lang': 'EN'
        }
        
        if dimension_filters:
            # Construct filter string in Eurostat format
            for dim, values in dimension_filters.items():
                if isinstance(values, list):
                    values_str = '+'.join(values)
                else:
                    values_str = values
                url = url.replace(f"/all", f"/{values_str}", 1)
        
        print(f"Fetching data from Eurostat: {dataset_id}")
        
        try:
            response = requests.get(url, params=params)
            response.raise_for_status()  # Raise an exception for HTTP errors
            
            data = response.json()
            
            # Process Eurostat JSON structure
            # Extract dimensions
            dimensions = data.get('dimension', {})
            
            # Extract values and prepare for DataFrame
            values = data.get('value', {})
            
            # Convert nested structure to a flat one
            rows = []
            
            # Get dimension IDs and their values
            dim_ids = list(dimensions.keys())
            dim_values = [list(dimensions[dim_id]['category']['label'].keys()) for dim_id in dim_ids]
            
            # Generate all combinations of dimension values
            import itertools
            for indices in itertools.product(*[range(len(vals)) for vals in dim_values]):
                # Construct the key for this combination
                key = ".".join(str(idx) for idx in indices)
                
                # Only include if the key exists in values
                if key in values:
                    # Create a row with dimension values and the data value
                    row = {}
                    for i, dim_id in enumerate(dim_ids):
                        dim_code = dim_values[i][indices[i]]
                        dim_label = dimensions[dim_id]['category']['label'][dim_code]
                        row[dim_id] = dim_code
                        row[f"{dim_id}_label"] = dim_label
                    
                    row['value'] = values[key]
                    rows.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(rows)
            
            # Cache data
            if use_cache:
                self._save_to_cache(df, 'eurostat', dataset_id, dimension_filters)
            
            return df
        
        except Exception as e:
            print(f"Error fetching Eurostat data: {str(e)}")
            return pd.DataFrame()
    
    def get_worldbank_data(self, indicator, country='POL', start_year=2000, end_year=2023, use_cache=True):
        """
        Get data from World Bank.
        
        Parameters:
        -----------
        indicator : str
            World Bank indicator code (e.g., 'SP.POP.TOTL')
        country : str or list
            Country code(s) (default: 'POL' for Poland)
        start_year : int
            Start year
        end_year : int
            End year
        use_cache : bool
            Whether to use cached data if available
            
        Returns:
        --------
        pandas.DataFrame
            DataFrame with World Bank data
        """
        # Try to get from cache if enabled
        params = {
            'country': country,
            'start_year': start_year,
            'end_year': end_year
        }
        
        if use_cache:
            cached_data = self._get_cached_data('worldbank', indicator, params)
            if cached_data is not None:
                return cached_data
        
        # Construct URL
        if isinstance(country, list):
            country_str = ';'.join(country)
        else:
            country_str = country
            
        url = f"{self.worldbank_base_url}country/{country_str}/indicator/{indicator}"
        request_params = {
            'date': f"{start_year}:{end_year}",
            'format': 'json',
            'per_page': 1000
        }
        
        print(f"Fetching data from World Bank: {indicator}")
        
        try:
            response = requests.get(url, params=request_params)
            response.raise_for_status()  # Raise an exception for HTTP errors
            
            data = response.json()
            
            # World Bank API returns a list where the first element is metadata
            # and the second element is the actual data
            if len(data) < 2 or not data[1]:
                print(f"No data returned for {indicator}")
                return pd.DataFrame()
            
            # Extract data
            rows = []
            for item in data[1]:
                row = {
                    'country': item.get('country', {}).get('value'),
                    'country_code': item.get('countryiso3code'),
                    'indicator': item.get('indicator', {}).get('value'),
                    'indicator_code': item.get('indicator', {}).get('id'),
                    'year': item.get('date'),
                    'value': item.get('value')
                }
                rows.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(rows)
            
            # Cache data
            if use_cache:
                self._save_to_cache(df, 'worldbank', indicator, params)
            
            return df
        
        except Exception as e:
            print(f"Error fetching World Bank data: {str(e)}")
            return pd.DataFrame()
    
    def append_to_input_file(self, df, sheet_name, start_cell='A1', overwrite_sheet=False):
        """
        Append or overwrite data in the input Excel file.
        
        Parameters:
        -----------
        df : pandas.DataFrame
            DataFrame to append
        sheet_name : str
            Sheet name to add or modify
        start_cell : str
            Starting cell for data (e.g., 'A1')
        overwrite_sheet : bool
            Whether to overwrite the existing sheet (if it exists)
            
        Returns:
        --------
        bool
            Success status
        """
        try:
            # Make backup of input file
            backup_file = f"{self.input_file}.bak"
            try:
                import shutil
                shutil.copy2(self.input_file, backup_file)
                print(f"Created backup of input file: {backup_file}")
            except Exception as e:
                print(f"Warning: Could not create backup: {str(e)}")
            
            # Check if file exists
            if not os.path.exists(self.input_file):
                # Create a new Excel file
                with pd.ExcelWriter(self.input_file, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                return True
            
            # Check if sheet exists
            book = load_workbook(self.input_file)
            
            if sheet_name in book.sheetnames and not overwrite_sheet:
                # Append to existing sheet
                with pd.ExcelWriter(self.input_file, engine='openpyxl', mode='a') as writer:
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    
                    # Get the starting row and column
                    start_row = int(re.search(r'\d+', start_cell).group())
                    start_col = start_cell[0]
                    
                    # Write to the specified location
                    df.to_excel(writer, sheet_name=sheet_name, startrow=start_row-1, 
                               startcol=ord(start_col)-ord('A'), index=False)
            else:
                # Add new sheet or overwrite existing
                with pd.ExcelWriter(self.input_file, engine='openpyxl', mode='a') as writer:
                    writer.book = book
                    
                    # If overwriting, remove existing sheet
                    if sheet_name in book.sheetnames:
                        std = book.get_sheet_by_name(sheet_name)
                        book.remove(std)
                    
                    # Write the new sheet
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            return True
            
        except Exception as e:
            print(f"Error appending to input file: {str(e)}")
            return False
    
    def process_population_data(self, country='PL', year=2022):
        """
        Process population data from Eurostat and add it to the input file.
        
        Parameters:
        -----------
        country : str
            Country code (e.g., 'PL' for Poland)
        year : int
            Year to retrieve
            
        Returns:
        --------
        pandas.DataFrame
            Processed population data
        """
        try:
            # Get population data from Eurostat
            population_data = self.get_eurostat_data(
                'demo_pjan',
                dimension_filters={'geo': country}
            )
            
            if population_data.empty:
                print("No population data retrieved")
                return pd.DataFrame()
            
            # Filter by year
            year_str = str(year)
            if 'time' in population_data.columns:
                population_data = population_data[population_data['time'] == year_str]
            
            # Process by age group and sex
            if 'age' in population_data.columns and 'sex' in population_data.columns:
                # Map sex codes to match our format
                sex_map = {'M': 'M', 'F': 'K', 'T': 'T'}
                if 'sex' in population_data.columns:
                    population_data['sex_mapped'] = population_data['sex'].map(sex_map)
                
                # Group by age and sex
                result = []
                for (age, sex), group in population_data.groupby(['age', 'sex_mapped']):
                    # Skip total
                    if age == 'TOTAL' or sex == 'T':
                        continue
                    
                    # Map age to our format (5-year groups)
                    age_num = None
                    if age.startswith('Y'):
                        try:
                            age_num = int(age[1:])
                            if age_num < 5:
                                age_group = '0-4'
                            elif age_num < 10:
                                age_group = '5-9'
                            elif age_num < 15:
                                age_group = '10-14'
                            elif age_num < 20:
                                age_group = '15-19'
                            elif age_num < 25:
                                age_group = '20-24'
                            elif age_num < 30:
                                age_group = '25-29'
                            elif age_num < 35:
                                age_group = '30-34'
                            elif age_num < 40:
                                age_group = '35-39'
                            elif age_num < 45:
                                age_group = '40-44'
                            elif age_num < 50:
                                age_group = '45-49'
                            elif age_num < 55:
                                age_group = '50-54'
                            elif age_num < 60:
                                age_group = '55-59'
                            elif age_num < 65:
                                age_group = '60-64'
                            elif age_num < 70:
                                age_group = '65-69'
                            elif age_num < 75:
                                age_group = '70-74'
                            elif age_num < 80:
                                age_group = '75-79'
                            elif age_num < 85:
                                age_group = '80-84'
                            else:
                                age_group = '85+'
                        except ValueError:
                            continue
                    elif age.startswith('Y_GE'):
                        # Y_GE85 means 85+
                        age_group = '85+'
                    else:
                        continue
                    
                    # Sum population for this age group and sex
                    population = group['value'].sum()
                    
                    result.append({
                        'year': year,
                        'age': age_group,
                        'sex': sex,
                        'population': population
                    })
                
                # Create DataFrame
                result_df = pd.DataFrame(result)
                
                # Add to input file (optional)
                if not result_df.empty:
                    self.append_to_input_file(
                        result_df,
                        sheet_name=f'population_{year}',
                        overwrite_sheet=True
                    )
                
                return result_df
            else:
                print("Required columns not found in population data")
                return pd.DataFrame()
            
        except Exception as e:
            print(f"Error processing population data: {str(e)}")
            return pd.DataFrame()
    
    def process_health_indicators(self, indicators, country='POL', start_year=2000, end_year=2022):
        """
        Process health indicators from World Bank and add them to the input file.
        
        Parameters:
        -----------
        indicators : list
            List of World Bank indicator codes
        country : str
            Country code (e.g., 'POL' for Poland)
        start_year : int
            Start year
        end_year : int
            End year
            
        Returns:
        --------
        pandas.DataFrame
            Processed health indicator data
        """
        try:
            # Get data for each indicator
            all_data = []
            
            for indicator in indicators:
                data = self.get_worldbank_data(
                    indicator,
                    country=country,
                    start_year=start_year,
                    end_year=end_year
                )
                
                if not data.empty:
                    all_data.append(data)
            
            if not all_data:
                print("No health indicator data retrieved")
                return pd.DataFrame()
            
            # Combine all indicators
            combined_data = pd.concat(all_data, ignore_index=True)
            
            # Pivot to have years as rows and indicators as columns
            pivoted = combined_data.pivot_table(
                index='year', 
                columns='indicator_code',
                values='value'
            ).reset_index()
            
            # Rename columns to more readable names
            readable_names = {
                'SP.DYN.LE00.IN': 'life_expectancy',
                'SH.DYN.MORT': 'under5_mortality',
                'SH.STA.DIAB.ZS': 'diabetes_prevalence',
                'SH.PRV.SMOK': 'smoking_prevalence',
                'SH.MED.BEDS.ZS': 'hospital_beds'
            }
            
            # Rename columns that exist in the data
            for code, name in readable_names.items():
                if code in pivoted.columns:
                    pivoted = pivoted.rename(columns={code: name})
            
            # Add to input file
            if not pivoted.empty:
                self.append_to_input_file(
                    pivoted,
                    sheet_name='health_indicators',
                    overwrite_sheet=True
                )
            
            return pivoted
            
        except Exception as e:
            print(f"Error processing health indicators: {str(e)}")
            return pd.DataFrame()


def download_and_prepare_data(country='PL', year=2022, output_file='Input.xlsx'):
    """
    Download and prepare data for simulation.
    
    Parameters:
    -----------
    country : str
        Country code for Eurostat (e.g., 'PL' for Poland)
    year : int
        Year for population data
    output_file : str
        Output file path
        
    Returns:
    --------
    bool
        Success status
    """
    try:
        # Initialize data integrator
        integrator = DataIntegrator(input_file=output_file)
        
        # Process population data
        print(f"Downloading population data for {country} ({year})...")
        population_df = integrator.process_population_data(country=country, year=year)
        
        if population_df.empty:
            print("Warning: No population data retrieved")
        else:
            print(f"Retrieved population data for {len(population_df)} demographic groups")
        
        # Process health indicators
        health_indicators = [
            'SP.DYN.LE00.IN',  # Life expectancy
            'SH.DYN.MORT',     # Under-5 mortality
            'SH.STA.DIAB.ZS',  # Diabetes prevalence
            'SH.PRV.SMOK',     # Smoking prevalence
            'SH.MED.BEDS.ZS'   # Hospital beds
        ]
        
        country_wb = 'POL' if country == 'PL' else country
        
        print(f"Downloading health indicators for {country_wb}...")
        health_df = integrator.process_health_indicators(
            indicators=health_indicators,
            country=country_wb,
            start_year=year-10,
            end_year=year
        )
        
        if health_df.empty:
            print("Warning: No health indicator data retrieved")
        else:
            print(f"Retrieved health indicator data for {len(health_df)} years")
        
        return not (population_df.empty and health_df.empty)
        
    except Exception as e:
        print(f"Error downloading and preparing data: {str(e)}")
        return False


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Download and prepare data for simulation')
    parser.add_argument('--country', type=str, default='PL', help='Country code (default: PL)')
    parser.add_argument('--year', type=int, default=2022, help='Year for population data (default: 2022)')
    parser.add_argument('--output', type=str, default='External_Input.xlsx', 
                       help='Output file path (default: External_Input.xlsx)')
    
    args = parser.parse_args()
    
    success = download_and_prepare_data(
        country=args.country,
        year=args.year,
        output_file=args.output
    )
    
    if success:
        print(f"Data successfully prepared and saved to {args.output}")
    else:
        print("Data preparation failed")
